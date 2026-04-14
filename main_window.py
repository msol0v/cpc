import faulthandler
faulthandler.enable()

from PyQt5 import QtWidgets, QtCore
from PyQt5.QtCore import QThread, pyqtSlot, pyqtSignal, QTimer
from PyQt5.QtWidgets import QRadioButton, QDesktopWidget, QTableWidgetItem, QDialog, QVBoxLayout, QLabel, QMessageBox
from PyQt5.QtGui import QColor, QBrush
from _ui_main_window import Ui_MainWindow
import os
import json
from arinc_handler import ArincWorker
from RS_handler import RsWorker
import BITE
from openpyxl import load_workbook
import uut_encode

base_dir = os.path.dirname(os.path.abspath(__file__))

class MainWindow(QtWidgets.QMainWindow):

    sig_run_arinc_handler = pyqtSignal()
    sig_change_ofv = pyqtSignal(bool, bool, int)

    sig_r_nvram = pyqtSignal(bool, int)
    sig_w_nvram = pyqtSignal(bool, int, int)

    sig_change_uut_in_words = pyqtSignal(dict)

    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        # Get com-ports names
        with open('ports.json', 'r', encoding='utf-8') as f:
            ports = json.load(f)

        # Treads
        self.arinc_thread = QThread()
        self.arinc_handler = ArincWorker(ports['arinc-com'])
        self.arinc_handler.moveToThread(self.arinc_thread)

        self.rs_thread = QThread()
        self.rs_handler = RsWorker()
        self.rs_handler.moveToThread(self.rs_thread)
        self.rs_thread.started.connect(self.rs_handler.start)


        # Signals connection
        self.arinc_handler.sig_error.connect(self.slot_arinc_handler_error)
        self.arinc_handler.sig_connected.connect(self.slot_arinc_handler_connected)
        self.sig_run_arinc_handler.connect(self.arinc_handler.slot_connect)

        self.ui.rButton.clicked.connect(self.on_rButton_clicked)
        self.sig_r_nvram.connect(self.rs_handler.slot_read)
        self.ui.wButton.clicked.connect(self.on_wButton_clicked)
        self.sig_w_nvram.connect(self.rs_handler.slot_write)
        self.rs_handler.sig_rw_reply.connect(self.slot_rw_reply)

        self.rs_handler.handler.sig_normal_resp.connect(self.slot_show_commanded_pos)
        self.rs_handler.signal_pc_alt.connect(self.slot_show_uut_to_operator)
        self.ui.readConstsButton.clicked.connect(self.rs_handler.slot_get_consts)
        self.rs_handler.sig_send_consts.connect(self.slot_show_consts)
        self.rs_handler.sig_progress_read_buffs.connect(self.slot_show_progress_read_buffs)

        # Save button
        self.ui.saveButton.clicked.connect(self.slot_save)

        # Faults Buffers
        self.ui.readFBuffersButton.clicked.connect(self.rs_handler.read_fbuffs)
        self.rs_handler.sig_show_fbuffs.connect(self.show_fbuffs)

        # uut in
        self.sig_change_uut_in_words.connect(self.arinc_handler.slot_change_uut_word)
        self.ui.setDefaultButton.clicked.connect(self.slot_set_default_button)
        self.ui.pushButton_setCustom_cfds.clicked.connect(self.slot_change_cfds)
        self.ui.pushButton_setCustom_fms.clicked.connect(self.slot_change_fms)
        self.ui.pushButton_setCustom_adirs.clicked.connect(self.slot_change_adirs)
        self.default_uut_in_params = {
            '246': '1020',
            '210': '100',
            '125': '855254AA',
            '126': 'E000006A',
            '227': '1105C0E9',
            '260': '14CA140D',
            '301': '529F3883',
            '302': 'C99B1043',
            '303': '00010CC3',
            '351': '00002097',
            '256': '992'
        }
        self.ui.pushButton_setCustom_ofv.clicked.connect(self.slot_change_ofv)
        self.sig_change_ofv.connect(self.rs_handler.slot_bite_change)
        self.default_ofv_pos = 50

        #labels
        self.arinc_handler.sig_bus_activity.connect(self.slot_bus_activity)
        self.arinc_handler.sig_352.connect(self.slot_352_word)
        self.arinc_handler.sig_353.connect(self.slot_353_word)
        self.arinc_handler.sig_163.connect(self.slot_163_word)
        self.arinc_handler.sig_057.connect(self.slot_057_word)
        self.arinc_handler.sig_104.connect(self.slot_104_word)
        self.arinc_handler.sig_105.connect(self.slot_105_word)
        self.arinc_handler.sig_106.connect(self.slot_106_word)
        self.arinc_handler.sig_107.connect(self.slot_107_word)
        self.arinc_handler.sig_110.connect(self.slot_110_word)
        self.arinc_handler.sig_111.connect(self.slot_111_word)

        ###
        BITE.ConfigureFaultBits()
        # Настройка таблицы (делается один раз при инициализации)
        self.ui.tableWidget.cellDoubleClicked.connect(self.show_row_info)

        self.ui.tableWidget.setColumnCount(4)
        self.ui.tableWidget.setHorizontalHeaderLabels(["Fault Code", "Name", "Class", "Cause"])
        self.ui.tableWidget.setRowCount(len(BITE.Faults))
        self.ui.tableWidget.setWordWrap(True)
        self.ui.tableWidget.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)

        current_font = self.ui.tableWidget.font()
        current_font.setPointSize(13)
        current_font.setBold(True)
        self.ui.tableWidget.setFont(current_font)

        for row, fault in enumerate(BITE.Faults):
            for col, value in enumerate([fault.code, fault.name, fault.fault_class, fault.cause]):
                item = QTableWidgetItem(str(value))
                item.setBackground(QColor(255, 255, 255))  # белый фон по умолчанию
                self.ui.tableWidget.setItem(row, col, item)

        self.ui.tableWidget.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.ui.tableWidget.resizeRowsToContents()
        self.ui.tableWidget.verticalHeader().setVisible(False)
        ###

        ### Настройка таблицы констант
        self.ui.tableWidget_consts.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.ui.tableWidget_consts.resizeRowsToContents()
        self.ui.tableWidget_consts.verticalHeader().setVisible(False)

        ### Настройка fbuffs
        self.ui.tableWidget_FBuffers.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.ui.tableWidget_FBuffers.resizeRowsToContents()
        self.ui.tableWidget_FBuffers.verticalHeader().setVisible(False)

        # Выставляем значения uut in по умолчанию и отправляем arinc_handler чтобы тоже выставил
        self.slot_set_default_button()


        self.ui.lcdNumber_pc.setDigitCount(8)
        self.ui.lcdNumber_alt.setDigitCount(8)

        self.arinc_thread.start()
        self.sig_run_arinc_handler.emit()
        self.rs_thread.start()

    @pyqtSlot()
    def slot_save(self):
        if self.ui.progressBar.value() == 100:
            if self.rs_handler.pn is None: self.rs_handler.pn = self.rs_handler.read_pn()
            if self.rs_handler.sn is None:  self.rs_handler.sn = self.rs_handler.read_sn()
            book = load_workbook(base_dir + '/base_xlsx/moduleCPC.xlsx')
            wl = book.get_sheet_by_name('com')
            for i, rbyte in enumerate(self.rs_handler.raw_buffs, 1):
                wl.cell(i,1,value=rbyte)
            book.save(f'{base_dir}/base_xlsx/moduleCPC_{self.rs_handler.pn}_{self.rs_handler.sn}.xlsx')


    @pyqtSlot()
    def slot_change_cfds(self):
        words = {
            '125': self.ui.lineEdit_125.text(),
            '126': self.ui.lineEdit_126.text(),
            '227': self.ui.lineEdit_227.text(),
            '260': self.ui.lineEdit_260.text(),
            '301': self.ui.lineEdit_301.text(),
            '302': self.ui.lineEdit_302.text(),
            '303': self.ui.lineEdit_303.text(),
        }
        self.sig_change_uut_in_words.emit(words)

    @pyqtSlot()
    def slot_change_fms(self):
        words = {
            '351': self.ui.valid_word_discrete.text(),
            '256': uut_encode.encode_256(int(self.ui.lfe_in_feet.text()))
        }
        self.sig_change_uut_in_words.emit(words)

    @pyqtSlot()
    def slot_change_adirs(self):
        words = {
            '246': uut_encode.encode_246(int(self.ui.static_pressure.text())),
            '210': uut_encode.encode_210(int(self.ui.true_airspeed.text())),
        }
        self.sig_change_uut_in_words.emit(words)

    @pyqtSlot()
    def slot_change_ofv(self):
        state = bool(self.ui.comboBox_ofv_state.currentIndex())
        bite_fault_flag = bool(self.ui.comboBox_ofv_bite.currentIndex())
        position = int(self.ui.lineEdit_ofv_input.text())
        if position < -2:
            position = -2
            self.ui.lineEdit_ofv_input.setText(f'{position}')
        if position > 122:
            position = 122
            self.ui.lineEdit_ofv_input.setText(f'{position}')
        self.sig_change_ofv.emit(state, bite_fault_flag, position)

    @pyqtSlot()
    def on_rButton_clicked(self):
        try:
            addr = int(self.ui.lineEdit_r_addr.text(), 16)
            if 0 > addr > 0xFFFF:
                print('Invalid address')
                self.ui.lineEdit_r_addr.setText("")
                return
        except ValueError:
            print('Invalid address')
            self.ui.lineEdit_r_addr.setText("")
            return
        state = self.ui.radioButton_r_0.isChecked()
        self.sig_r_nvram.emit(state, addr)


    @pyqtSlot()
    def on_wButton_clicked(self):
        try:
            addr = int(self.ui.lineEdit_w_addr.text(), 16)
            data = int(self.ui.lineEdit_w_data.text(), 16)
            if 0 > addr > 0xFFFF:
                print('Invalid address')
                return
        except ValueError:
            print('Invalid address')
            return
        state = self.ui.radioButton_w_0.isChecked()
        self.sig_w_nvram.emit(state, addr, data)

    @pyqtSlot(int, str)
    def slot_rw_reply(self, r0_w1, reply):
        if r0_w1 == 0:
            self.ui.lineEdit_r_reply.setText(reply)
        else:
            self.ui.lineEdit_w_reply.setText(reply)

    @pyqtSlot(int, int)
    def show_row_info(self, row, column):
        dlg = QDialog(self)
        dlg.setWindowTitle(f'Info {BITE.Faults[row].name}')
        dlg.setMinimumWidth(300)
        layout = QVBoxLayout()
        label = QLabel(BITE.Faults[row].descriptor)
        layout.addWidget(label)
        dlg.setLayout(layout)
        dlg.show()

    def parse_dict_to_radio(self, dict_data):
        for key, value in dict_data.items():
            attr_1, attr_0 = (key + '_1', key + '_0')
            if hasattr(self.ui, attr_1) and hasattr(self.ui, attr_0):
                box_1: QRadioButton = getattr(self.ui, attr_1)
                box_0: QRadioButton = getattr(self.ui, attr_0)
                if value:
                    box_1.setChecked(True)
                else:
                    box_0.setChecked(True)

    def set_row_color(self,fault_bits, base):
        for bit in range(16):
            bit_value = (fault_bits >> bit) & 1
            color = QColor(255, 165, 0) if bit_value else QColor(255, 255, 255)
            for col in range(self.ui.tableWidget.columnCount()):
                item = self.ui.tableWidget.item(bit+base, col)
                if item:
                    item.setBackground(QBrush(color))

        self.ui.tableWidget.viewport().update()

    @pyqtSlot(int)
    def slot_show_progress_read_buffs(self, progr: int):
        self.ui.progressBar.setValue(progr)

    @pyqtSlot(list)
    def show_fbuffs(self, buffs: list):
        self.ui.tableWidget_FBuffers.setRowCount(len(buffs))
        for i, buffer in enumerate(buffs):
            for j, (key, value) in enumerate(buffer.__dict__.items()):
                self.ui.tableWidget_FBuffers.setItem(i, j, QTableWidgetItem(str(value)))

    @pyqtSlot(dict)
    def slot_bus_activity(self, active):
        if active.get('sdac_0'):
            self.ui.sdac_1.setChecked(True)
        else:
            self.ui.sdac_1.setChecked(False)

        if active.get('sdac_1'):
            self.ui.sdac_2.setChecked(True)
        else:
            self.ui.sdac_2.setChecked(False)


    @pyqtSlot()
    def slot_set_default_button(self):
        self.ui.lineEdit_125.setText(self.default_uut_in_params['125'])
        self.ui.lineEdit_126.setText(self.default_uut_in_params['126'])
        self.ui.lineEdit_227.setText(self.default_uut_in_params['227'])
        self.ui.lineEdit_260.setText(self.default_uut_in_params['260'])
        self.ui.lineEdit_301.setText(self.default_uut_in_params['301'])
        self.ui.lineEdit_302.setText(self.default_uut_in_params['302'])
        self.ui.lineEdit_303.setText(self.default_uut_in_params['303'])
        self.ui.valid_word_discrete.setText(self.default_uut_in_params['351'])
        self.ui.lfe_in_feet.setText(self.default_uut_in_params['256'])
        self.ui.static_pressure.setText(self.default_uut_in_params['246'])
        self.ui.true_airspeed.setText(self.default_uut_in_params['210'])

        self.ui.lineEdit_ofv_input.setText(f'{self.default_ofv_pos}')
        self.ui.comboBox_ofv_state.setCurrentIndex(1) #Operational
        self.ui.comboBox_ofv_bite.setCurrentIndex(0)

        self.slot_change_fms()
        self.slot_change_cfds()
        self.slot_change_adirs()
        self.slot_change_ofv()


    @pyqtSlot(dict)
    def slot_show_consts(self, consts: dict):
        self.ui.tableWidget_consts.setRowCount(len(consts))
        for i, (key, value) in enumerate(consts.items()):
            self.ui.tableWidget_consts.setItem(i, 0, QTableWidgetItem(key))
            self.ui.tableWidget_consts.setItem(i, 1, QTableWidgetItem(str(value)))

    @pyqtSlot(tuple)
    def slot_show_commanded_pos(self, br_zc):
        br, zc = br_zc
        self.ui.lcdNumber_comPos.display(br)

    @pyqtSlot(tuple)
    def slot_show_uut_to_operator(self, pc_alt):
        pc, alt = pc_alt
        self.ui.lcdNumber_pc.display(pc)
        self.ui.lcdNumber_alt.display(alt)

    @pyqtSlot(dict)
    def slot_352_word(self, disc_sig_dict):
        if disc_sig_dict['FAULT_WARN_WR']:
            self.ui.FAULT_WARN_WR_1.setChecked(True)
        else:
            self.ui.FAULT_WARN_WR_0.setChecked(True)

        if disc_sig_dict['SYS_IN_CNTL_IN']:
            self.ui.SYS_IN_CNTL_IN_1.setChecked(True)
        else:
            self.ui.SYS_IN_CNTL_IN_0.setChecked(True)

        if disc_sig_dict['ENG_1_N2']:
            self.ui.ENG1_N2_1.setChecked(True)
        else:
            self.ui.ENG1_N2_0.setChecked(True)

        if disc_sig_dict['ENG_2_N2']:
            self.ui.ENG2_N2_1.setChecked(True)
        else:
            self.ui.ENG2_N2_0.setChecked(True)

        if disc_sig_dict['LDG_GEAR_NORM']:
            self.ui.LDG_GEAR_NORM_1.setChecked(True)
        else:
            self.ui.LDG_GEAR_NORM_0.setChecked(True)

        if disc_sig_dict['LDG_GEAR_ESS']:
            self.ui.LDG_GEAR_ESS_1.setChecked(True)
        else:
            self.ui.LDG_GEAR_ESS_0.setChecked(True)

    @pyqtSlot(dict)
    def slot_353_word(self, disc_sig_dict):
        self.parse_dict_to_radio(disc_sig_dict)

    @pyqtSlot(int)
    def slot_163_word(self, data):
        self.ui.lcdNumber.display(data)

    @pyqtSlot(dict)
    def slot_057_word(self, data):
        if data['sys_in_control']:
            self.ui.sys_in_ctrl_yes.setChecked(True)
        else:
            self.ui.sys_in_ctrl_no.setChecked(True)

        if data['sys_status']:
            self.ui.sys_status_fail.setChecked(True)
        else:
            self.ui.sys_status_good.setChecked(True)

        if data['lfe_from_fms']:
            self.ui.lfe_from_fms_nused.setChecked(True)
        else:
            self.ui.lfe_from_fms_used.setChecked(True)

        if data['excessive_cabin_altitude']:
            self.ui.cabine_altitude_warn.setChecked(True)
        else:
            self.ui.cabine_altitude_nwarn.setChecked(True)

        if data['low_differential_pressure']:
            self.ui.diff_pressure_warn.setChecked(True)
        else:
            self.ui.diff_pressure_nwarn.setChecked(True)

        if data['preplanned_descent_info']:
            self.ui.preplanned_toofast.setChecked(True)
        else:
            self.ui.preplanned_good.setChecked(True)

        if data['lfes_status']:
            self.ui.lfes_stat_manual.setChecked(True)
        else:
            self.ui.lfes_stat_auto.setChecked(True)

        val = data['used_adirs_channel']
        chan = (val & 0x02) >> 1 | (val & 0x01) << 1
        match chan:
            case 0:
                self.ui.adirs_ch_no.setChecked(True)
            case 1:
                self.ui.adirs_ch_1.setChecked(True)
            case 2:
                self.ui.adirs_ch_2.setChecked(True)
            case 3:
                self.ui.adirs_ch_3.setChecked(True)

        if data['fms_enable']:
            self.ui.fms_enable.setChecked(True)
        else:
            self.ui.fms_disable.setChecked(True)


        used_mode = data['flight_modes']
        used_mode = (used_mode & 0x01) << 2 | (used_mode & 0x02) | (used_mode & 0x04) >> 2
        modes = ['gr', 'to', 'ci', 'ce', 'cr', 'di', 'de', 'ab']
        for i, mode in enumerate(modes):
            widget: QRadioButton = getattr(self.ui, f'flight_modes_{mode}')
            if i == used_mode:
                widget.setChecked(True)

        val = data['fms_selection']
        fms_sel = (val & 0x02) >> 1 | (val & 0x01) << 1
        match fms_sel:
            case 0:
                self.ui.fms_sel_1eng.setChecked(True)
            case 1:
                self.ui.fms_sel_1neng.setChecked(True)
            case 2:
                self.ui.fms_sel_2eng.setChecked(True)
            case 3:
                self.ui.fms_sel_2neng.setChecked(True)

        if data['fms_use']:
            self.ui.fms_use_nuse.setChecked(True)
        else:
            self.ui.fms_use_use.setChecked(True)

        if data['qnh_from_fms']:
            self.ui.qnh_nuse.setChecked(True)
        else:
            self.ui.qnh_use.setChecked(True)

        if data['class_2_fault']:
            self.ui.class2fault_pres.setChecked(True)
        else:
            self.ui.class2fault_npres.setChecked(True)

    @pyqtSlot()
    def slot_arinc_handler_connected(self):
        print('arinc handler connected')

    @pyqtSlot(str)
    def slot_arinc_handler_error(self, error:str):
        QMessageBox.critical(self, "ARINC Handler Error", error)

    @pyqtSlot(int)
    def slot_104_word(self, fault_bits: int):
        self.set_row_color(fault_bits, BITE.base[104])

    @pyqtSlot(int)
    def slot_105_word(self, fault_bits: int):
        self.set_row_color(fault_bits, BITE.base[105])

    @pyqtSlot(int)
    def slot_106_word(self, fault_bits: int):
        self.set_row_color(fault_bits, BITE.base[106])

    @pyqtSlot(int)
    def slot_107_word(self, fault_bits: int):
        self.set_row_color(fault_bits, BITE.base[107])

    @pyqtSlot(int)
    def slot_110_word(self, fault_bits: int):
        self.set_row_color(fault_bits, BITE.base[110])

    @pyqtSlot(int)
    def slot_111_word(self, fault_bits: int):
        self.set_row_color(fault_bits, BITE.base[111])

